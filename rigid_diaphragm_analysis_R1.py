"""
Rigid diaphragm (plan) shear distribution script
================================================

This script reproduces the calculation logic in the provided Excel template:
    Rigid_only_test.xlsm  (sheet: "Rigid_Template (2)")

Inputs requested from the user match:
    - Blue cells J10:J29 (only the populated ones: J10, J11, J15, J16, J17, J19, J20, J28, J29)
    - Wall table D49:I52 (Wall Name, Length, Height, w_k, x_coord, y_coord)

It computes:
    - Center of Mass (CoM) in the sheet's working coordinate system
    - Center of Rigidity (CoR)
    - Real eccentricity ex, ey
    - Polar "torsional rigidity" term Jp (with corrected V = Rix*ybar^2 term)
    - Direct shear distribution
    - Torsional shear from real eccentricity
    - Torsional shear from 10% accidental eccentricity
    - Total Vx and Vy per wall (matching the Excel table)

USAGE (Jupyter):
    from rigid_diaphragm_analysis import run_interactive, compute_rigid_diaphragm

    # 1) Interactive (uses defaults read from the template):
    df, summary = run_interactive(excel_path="Rigid_only_test.xlsm")
    df

    # 2) Programmatic:
    inputs = {...}
    walls = [...]
    df, summary = compute_rigid_diaphragm(inputs, walls)

OPTIONAL VALIDATION (for the template case):
    from rigid_diaphragm_analysis import validate_against_excel
    validate_against_excel("Rigid_only_test.xlsm")
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import math

import pandas as pd
import openpyxl


# -----------------------------
# Helpers
# -----------------------------

def _prompt_float(label: str, default: float) -> float:
    """Prompt for a float in console/Jupyter input(); press Enter to accept default."""
    raw = input(f"{label} [{default}]: ").strip()
    return float(raw) if raw else float(default)


def _clean_wall_name(name: str) -> str:
    return str(name).strip()


def _wall_orientation_factors(wall_name: str, Di: float) -> Tuple[float, float]:
    """
    Match the Excel mapping:
      - 'EW' walls: Rix = 1/Di, Riy = 0
      - 'NS' walls: Rix = 0, Riy = 1/Di

    NOTE: This is the spreadsheet's convention (it must be followed for matching results),
    even if a different naming convention is used elsewhere.
    """
    nm = wall_name.upper()
    if "EW" in nm:
        return (1.0 / Di, 0.0)
    if "NS" in nm:
        return (0.0, 1.0 / Di)
    raise ValueError(
        f"Wall Name '{wall_name}' must contain 'EW' or 'NS' so the script can assign Rix/Riy."
    )


# -----------------------------
# Core calculation
# -----------------------------

def compute_rigid_diaphragm(
    inputs: Dict[str, float],
    walls: List[Dict[str, object]],
) -> Tuple[pd.DataFrame, Dict[str, float]]:
    """
    Compute rigid diaphragm results matching the Excel template.

    Required inputs keys:
        L, B, Xcm_manual, Ycm_manual, W, X_offset, Y_offset, Fx, Fy

    Required wall dict keys:
        "Wall Name", "Length (m)", "Height (m)", "w_k (kN)", "x_coord (m)", "y_coord (m)"
    """
    # Unpack
    L = float(inputs["L"])
    B = float(inputs["B"])
    Xcm_manual = float(inputs["Xcm_manual"])
    Ycm_manual = float(inputs["Ycm_manual"])
    W = float(inputs["W"])  # used for completeness; current template doesn't use it downstream
    X_offset = float(inputs["X_offset"])
    Y_offset = float(inputs["Y_offset"])
    Fx = float(inputs["Fx"])
    Fy = float(inputs["Fy"])

    # Accidental eccentricities (Excel: exbar=0.1*L, eybar=0.1*B)
    exbar = 0.1 * L
    eybar = 0.1 * B

    # 1) Transform to the working coordinate system:
    #    Excel uses "paper offsets" so that a specific paper coordinate becomes (0,0).
    #    xk = x_coord - X_offset, yk = y_coord - Y_offset
    #    Xcm = Xcm_manual - X_offset, Ycm = Ycm_manual - Y_offset
    Xcm = Xcm_manual - X_offset
    Ycm = Ycm_manual - Y_offset

    rows = []
    for w in walls:
        name = _clean_wall_name(w["Wall Name"])
        length = float(w["Length (m)"])
        height = float(w["Height (m)"])
        wk = w.get("w_k (kN)", None)
        x_coord = float(w["x_coord (m)"])
        y_coord = float(w["y_coord (m)"])

        xk = x_coord - X_offset
        yk = y_coord - Y_offset

        # 2) Di and rigidity factors (exact Excel expression)
        #    Di = 4*(h/L)^3 + 3*(h/L)
        Di = 4.0 * (height / length) ** 3 + 3.0 * (height / length)

        # 3) Directional rigidities per spreadsheet convention
        Rix, Riy = _wall_orientation_factors(name, Di)

        rows.append(
            {
                "Wall Name": name,
                "Length (m)": length,
                "Height (m)": height,
                "w_k (kN)": wk,
                "x_coord (m)": x_coord,
                "y_coord (m)": y_coord,
                "xk (m)": xk,
                "yk (m)": yk,
                "Di": Di,
                "Rix": Rix,
                "Riy": Riy,
                "Riy*xk": Riy * xk,
                "Rix*yk": Rix * yk,  # Excel column R is "Rix*yk"
            }
        )

    df = pd.DataFrame(rows)

    sum_Rix = float(df["Rix"].sum())
    sum_Riy = float(df["Riy"].sum())

    # 4) Center of Rigidity (Excel formulas):
    #    Xcr = SUM(Riy*xk) / SUM(Riy)
    #    Ycr = SUM(Rix*yk) / SUM(Rix)
    Xcr = (float(df["Riy*xk"].sum()) / sum_Riy) if sum_Riy != 0 else float("nan")
    Ycr = (float(df["Rix*yk"].sum()) / sum_Rix) if sum_Rix != 0 else float("nan")

    # 5) Real eccentricities (absolute values per Excel)
    ex = abs(Xcm - Xcr)
    ey = abs(Ycm - Ycr)

    # 6) Relative distances to CoR
    df["xbar"] = df["xk (m)"] - Xcr
    df["ybar"] = df["yk (m)"] - Ycr

    # 7) Jp (match Excel computation):
    #    U = Riy*xbar^2
    #    V = (Rix)*ybar^2      <-- corrected per user (no extra *yk factor)
    #    Jp = SUM(U) + SUM(V)
    df["Riy*xbar2"] = df["Riy"] * (df["xbar"] ** 2)
    df["Rix*ybar2"] = df["Rix"] * (df["ybar"] ** 2)
    Jp = float(df["Riy*xbar2"].sum() + df["Rix*ybar2"].sum())

    # 8) Shear from torsion caused by real eccentricity (Excel columns W-Z)
    df["Real Tor Ratio_x"] = df["Rix"] * df["ybar"] * ey / Jp
    df["Real Tor Ratio_y"] = df["Riy"] * df["xbar"] * ex / Jp
    df["Vx_Real Tor"] = Fx * df["Real Tor Ratio_x"]
    df["Vy_Real Tor"] = Fy * df["Real Tor Ratio_y"]

    # 9) Direct shear due to force applied (Excel columns AA-AD)
    df["Direct Shear Ratio_x"] = (df["Rix"] / sum_Rix) if sum_Rix != 0 else 0.0
    df["Direct Shear Ratio_y"] = (df["Riy"] / sum_Riy) if sum_Riy != 0 else 0.0
    df["Direct Shear_x"] = Fx * df["Direct Shear Ratio_x"]
    df["Direct Shear_y"] = Fy * df["Direct Shear Ratio_y"]

    # 10) Shear from torsion caused by 10% accidental eccentricity (Excel columns AE-AH)
    #     NOTE: This block reproduces the spreadsheet's exact multiplications:
    #       AccTorRatio_x = eybar*Rix*ybar/Jp  and Vx_Acc_Tor = Fy*AccTorRatio_x
    #       AccTorRatio_y = exbar*Riy*xbar/Jp  and Vy_Acc_Tor = Fx*AccTorRatio_y
    df["AccTorRatio_x"] = eybar * df["Rix"] * df["ybar"] / Jp
    df["AccTorRatio_y"] = exbar * df["Riy"] * df["xbar"] / Jp
    df["Vx_Acc_Tor"] = Fy * df["AccTorRatio_x"]
    df["Vy_Acc_Tor"] = Fx * df["AccTorRatio_y"]

    # 11) Total shear (Excel columns AI-AJ)
    df["Vx (kN)"] = df["Vx_Real Tor"] + df["Direct Shear_x"] + df["Vx_Acc_Tor"]
    df["Vy (kN)"] = df["Vy_Real Tor"] + df["Direct Shear_y"] + df["Vy_Acc_Tor"].abs()

    # Column ordering to mirror the Excel table (D ... AJ), as a DataFrame
    ordered_cols = [
        "Wall Name", "Length (m)", "Height (m)", "w_k (kN)",
        "x_coord (m)", "y_coord (m)", "xk (m)", "yk (m)",
        "Di", "Rix", "Riy", "Riy*xk", "Rix*yk",
        "xbar", "ybar", "Riy*xbar2", "Rix*ybar2",
        "Real Tor Ratio_x", "Real Tor Ratio_y", "Vx_Real Tor", "Vy_Real Tor",
        "Direct Shear Ratio_x", "Direct Shear Ratio_y", "Direct Shear_x", "Direct Shear_y",
        "AccTorRatio_x", "AccTorRatio_y", "Vx_Acc_Tor", "Vy_Acc_Tor",
        "Vx (kN)", "Vy (kN)",
    ]
    df = df[ordered_cols]

    summary = {
        "L": L,
        "B": B,
        "X_offset": X_offset,
        "Y_offset": Y_offset,
        "Xcm": Xcm,
        "Ycm": Ycm,
        "Xcr": Xcr,
        "Ycr": Ycr,
        "ex": ex,
        "ey": ey,
        "exbar": exbar,
        "eybar": eybar,
        "Jp": Jp,
        "Fx": Fx,
        "Fy": Fy,
        "sum_Rix": sum_Rix,
        "sum_Riy": sum_Riy,
    }
    return df, summary


# -----------------------------
# Reading defaults from Excel
# -----------------------------

def load_defaults_from_excel(excel_path: str, sheet_name: str = "Rigid_Template (2)") -> Tuple[Dict[str, float], List[Dict[str, object]]]:
    """
    Load default input values and wall table entries from the Excel template.

    Reads:
      - J10, J11, J15, J16, J17, J19, J20, J28, J29
      - D49:I52

    Uses data_only=True so the stored numeric values are returned.
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)
    ws = wb[sheet_name]

    inputs = {
        "L": float(ws["J10"].value),
        "B": float(ws["J11"].value),
        "Xcm_manual": float(ws["J15"].value),
        "Ycm_manual": float(ws["J16"].value),
        "W": float(ws["J17"].value),
        "X_offset": float(ws["J19"].value),
        "Y_offset": float(ws["J20"].value),
        "Fx": float(ws["J28"].value),
        "Fy": float(ws["J29"].value),
    }

    walls: List[Dict[str, object]] = []
    for r in range(49, 53):
        name = ws[f"D{r}"].value
        if name is None:
            continue
        walls.append(
            {
                "Wall Name": str(name).strip(),
                "Length (m)": float(ws[f"E{r}"].value),
                "Height (m)": float(ws[f"F{r}"].value),
                "w_k (kN)": ws[f"G{r}"].value,
                "x_coord (m)": float(ws[f"H{r}"].value),
                "y_coord (m)": float(ws[f"I{r}"].value),
            }
        )
    return inputs, walls


def run_interactive(excel_path: Optional[str] = None) -> Tuple[pd.DataFrame, Dict[str, float]]:
    """
    Interactive prompt for inputs. If excel_path is provided, values from the template are used as defaults.
    """
    if excel_path:
        defaults_inputs, defaults_walls = load_defaults_from_excel(excel_path)
    else:
        raise ValueError("Please provide excel_path so the script can show sensible defaults.")

    print("\n--- Diaphragm-level inputs (press Enter to accept default) ---")
    inputs = {}
    inputs["L"] = _prompt_float("L  (EW / X dimension) [m]", defaults_inputs["L"])
    inputs["B"] = _prompt_float("B  (NS / Y dimension) [m]", defaults_inputs["B"])
    inputs["Xcm_manual"] = _prompt_float("Xcm (manual, in paper coordinates) [m]", defaults_inputs["Xcm_manual"])
    inputs["Ycm_manual"] = _prompt_float("Ycm (manual, in paper coordinates) [m]", defaults_inputs["Ycm_manual"])
    inputs["W"] = _prompt_float("W  (diaphragm weight) [kN]", defaults_inputs["W"])
    inputs["X_offset"] = _prompt_float("X_paper_offset (coordinate to treat as x=0) [m]", defaults_inputs["X_offset"])
    inputs["Y_offset"] = _prompt_float("Y_paper_offset (coordinate to treat as y=0) [m]", defaults_inputs["Y_offset"])
    inputs["Fx"] = _prompt_float("F_x_dir (lateral force in x direction) [kN]", defaults_inputs["Fx"])
    inputs["Fy"] = _prompt_float("F_y_dir (lateral force in y direction) [kN]", defaults_inputs["Fy"])

    print("\n--- Wall table (press Enter to accept default) ---")
    walls: List[Dict[str, object]] = []
    for i, w in enumerate(defaults_walls, start=1):
        print(f"\nWall {i}:")
        name = input(f"  Wall Name [{w['Wall Name']}]: ").strip() or str(w["Wall Name"])
        length = _prompt_float("  Length (m)", float(w["Length (m)"]))
        height = _prompt_float("  Height (m)", float(w["Height (m)"]))
        # w_k is not used by the current template's downstream math, but we keep it
        wk_default = w["w_k (kN)"] if w["w_k (kN)"] is not None else 0.0
        wk_raw = input(f"  w_k (kN) [stored={wk_default}]: ").strip()
        wk = float(wk_raw) if wk_raw else wk_default
        x_coord = _prompt_float("  x_coord (m)", float(w["x_coord (m)"]))
        y_coord = _prompt_float("  y_coord (m)", float(w["y_coord (m)"]))

        walls.append(
            {
                "Wall Name": name,
                "Length (m)": length,
                "Height (m)": height,
                "w_k (kN)": wk,
                "x_coord (m)": x_coord,
                "y_coord (m)": y_coord,
            }
        )

    df, summary = compute_rigid_diaphragm(inputs, walls)

    print("\n--- Key results ---")
    for k in ["Xcm", "Ycm", "Xcr", "Ycr", "ex", "ey", "exbar", "eybar", "Jp"]:
        print(f"{k:>6s} = {summary[k]:.6g}")

    return df, summary


# -----------------------------
# Validation against the template
# -----------------------------

def validate_against_excel(excel_path: str, sheet_name: str = "Rigid_Template (2)", tol: float = 1e-6) -> pd.DataFrame:
    """
    Cross-check the script output against the stored Excel outputs for the provided template file.

    This validates the computation against the workbook's current stored results (data_only=True).
    """
    inputs, walls = load_defaults_from_excel(excel_path, sheet_name=sheet_name)
    df_calc, summary = compute_rigid_diaphragm(inputs, walls)

    wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=True)
    ws = wb[sheet_name]

    excel_rows = []
    for r in range(49, 53):
        name = str(ws[f"D{r}"].value).strip()
        excel_rows.append(
            {
                "Wall Name": name,
                "Excel Vx (kN)": float(ws[f"AI{r}"].value),
                "Excel Vy (kN)": float(ws[f"AJ{r}"].value),
                "Excel Vx_Real Tor": float(ws[f"Y{r}"].value),
                "Excel Vy_Real Tor": float(ws[f"Z{r}"].value),
                "Excel Direct Shear_x": float(ws[f"AC{r}"].value),
                "Excel Direct Shear_y": float(ws[f"AD{r}"].value),
                "Excel Vx_Acc_Tor": float(ws[f"AG{r}"].value),
                "Excel Vy_Acc_Tor": float(ws[f"AH{r}"].value),
            }
        )

    df_excel = pd.DataFrame(excel_rows)
    df_merge = df_calc.merge(df_excel, on="Wall Name", how="left")

    # Differences
    df_merge["dVx"] = df_merge["Vx (kN)"] - df_merge["Excel Vx (kN)"]
    df_merge["dVy"] = df_merge["Vy (kN)"] - df_merge["Excel Vy (kN)"]

    ok_vx = (df_merge["dVx"].abs() <= tol).all()
    ok_vy = (df_merge["dVy"].abs() <= tol).all()

    print("\nValidation summary:")
    print(f"  Jp (script) = {summary['Jp']:.12g}")
    print(f"  Vx match within tol={tol}: {ok_vx}")
    print(f"  Vy match within tol={tol}: {ok_vy}")

    return df_merge


if __name__ == "__main__":
    # If you run this as a script, it will ask you for an Excel path then prompt for inputs.
    excel_path = input("Excel file path (e.g., Rigid_only_test.xlsm): ").strip()
    df, summary = run_interactive(excel_path=excel_path)
    print("\n--- Output table ---")
    print(df.to_string(index=False))
